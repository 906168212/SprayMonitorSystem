from tkinter import *
from tkinter import messagebox
from tkinter.scrolledtext import ScrolledText
import serial
import serial.tools.list_ports
import threading
import queue
import datetime
import json
from openpyxl import Workbook


def get_local_version():
    try:
        with open('version.json', 'r') as f:
            return json.load(f)['latest_version']
    except FileNotFoundError:
        return '0.0.0'


class Concert(Frame):
    """初始化"""

    def __init__(self, master=None):
        super().__init__(master)
        # 1s定时器
        self.timer = None
        # 串口
        self.ser = None
        self.running_ser = False  # 串口运行标志
        self.port_list = ['']  # 串口列表
        self.baudrate_list = ['9600', '19200', '38400', '57600', '115200']
        self.port_var = StringVar()  # 串口Var句柄
        self.port_choose = None
        self.baudrate_var = StringVar()  # 串口波特率句柄
        self.baudrate_choose = None
        self.serial_button = None
        self.refresh_button = None
        self.flow_volumes = []
        self.channel_flow_var = StringVar()  # 管路流量句柄
        self.pressure_var = StringVar()  # 管路压力句柄
        self.set_freq_button = None
        self.set_duty_phase_button = None
        self.close_channel_button = None
        self.close_channel_window = None
        self.channels = [IntVar(),IntVar(),IntVar(),IntVar(),IntVar()]
        self.message_box = None
        self.send_queue = queue.Queue()  # 创建发送任务队列
        self.receive_thread = None  # 串口接收后台线程句柄
        self.send_thread = None  # 串口发送后台线程句柄
        self.serial_lock = threading.Lock()  # 创建串口线程锁，保证收发数据的正常
        # 自动发送读取流量CAN指令
        self.auto_read_flow_flag = True
        # 信息显示
        self.message_queue = queue.Queue()  # 窗口显示改为定时器批量显示，且放入消息队列中以此显示
        self.after_id = None
        self.MESSAGE_UPDATE_TIME = 200  # 200ms 更新一次界面
        # excel
        self.excel_data_queue = queue.Queue()  # excel数据存储队列，避免主线程阻塞
        self.saver_thread = None  # excel后台保存线程句柄
        self.running_saver = False  # 保存线程运行标志
        self.save_lock = threading.Lock()  # Excel保存锁
        self.xlsx_name = ''
        self.workbook = None
        self.worksheet = None
        # 主界面
        self.window_width = 900
        self.window_height = 600
        self.master = master
        self.pack()
        root.protocol("WM_DELETE_WINDOW", self.on_closing)  # 窗口关闭事件绑定
        self.create_widget()  # 界面组件创建
        # # excel初始化
        # self.excel_init()

    '''更新系统'''

    '''日志系统'''

    def excel_init(self):
        try:
            self.workbook = Workbook()  # 创建工作薄
            self.worksheet = self.workbook.active  # 选择默认工作表
            self.worksheet.merge_cells('B1:E1')  # 透传固定帧
            self.worksheet.merge_cells('F1:I1')  # CAN ID
            self.worksheet.merge_cells('J1:Q1')  # CAN 数据
            self.worksheet._current_row = 0
            self.worksheet.append(
                {1: '时间', 2: '透传固定帧', 6: 'CAN ID', 10: 'CAN数据'})  # ['时间','透传固定帧','CAN ID','CAN数据']
            self.xlsx_name = '试验CAN通信数据日志-' + str(
                datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')) + '.xlsx'
            self.workbook.save(self.xlsx_name)  # 保存工作薄
        except Exception as e:
            messagebox.showerror('Error', '初始化Excel时出错：' + str(e))

    def save_can_info(self, can_info):
        try:
            timestamp = datetime.datetime.now()  # 获取当前时间
            self.excel_data_queue.put((timestamp, can_info))
        except Exception as e:
            self.message_display(f"保存数据到队列失败：{str(e)}", 'red')

    '''excel 后台保存线程'''

    def saver_thread_func(self):
        buffer = []
        buffer_max = 100
        while self.running_saver or not self.excel_data_queue.empty():
            try:
                # 非阻塞获取数据，最多等待1秒
                item = self.excel_data_queue.get(timeout=1)
                timestamp, can_info = item
                buffer.append([timestamp] + can_info)
                if len(buffer) >= buffer_max:
                    self._save_to_excel(buffer)
                    buffer = []
            except queue.Empty:
                # 超时后检查是否需要退出
                if not self.running_saver:
                    break
            except Exception as e:
                self.message_display(f'后台保存出错：{str(e)}', 'red')
        # 退出前保存剩余数据
        if buffer:
            self.message_display('串口已关闭，进行剩余数据保存...', 'blue')
            self._save_to_excel(buffer)

    '''私有方法执行实际保存操作'''

    def _save_to_excel(self, data_buffer):
        try:
            with self.save_lock:  # 锁住Excel线程，保证线程安全
                for row in data_buffer:
                    self.worksheet.append(row)
                self.workbook.save(self.xlsx_name)  # 保存
                self.message_display(f'成功保存{len(data_buffer)}条数据', 'green')
        except PermissionError:
            self.message_display(f"Excel文件被占用，请关闭已打开的文件！", 'red')
        except Exception as e:
            self.message_display(f"Excel后台保存数据时出错：{str(e)}", 'red')

    '''窗口关闭时执行一些操作'''

    def on_closing(self):
        # 关闭串口，已经关闭则无需再关闭
        if self.running_ser:
            self.switch_serial_state()
        if not self.excel_data_queue.empty():
            self.message_display("正在保存剩余数据...", "blue")
            self.running_saver = True  # 临时恢复标志让线程处理
            self.saver_thread = threading.Thread(target=self.saver_thread_func(), daemon=True)
            self.saver_thread.start()
            self.saver_thread.join(timeout=10)  # 最多等待10秒
            # 其他关闭操作
        # 确保窗口被正确关闭
        root.destroy()

    '''接收线程'''

    def receive_thread_func(self):
        buffer = b''  # 数据缓存
        while self.running_ser:
            try:
                with self.serial_lock:  # 锁住串口线程，保证线程安全
                    if self.ser and self.ser.in_waiting >= 16:  # 至少读取一个完整帧
                        buffer += self.ser.read(16 * (self.ser.in_waiting // 16))
                while len(buffer) >= 16:
                    if buffer[0] == 0xAA:
                        message = buffer[:16]  # 提取完整信息
                        buffer = buffer[16:]  # 移除已处理部分
                        # 转为16进制字符串
                        hex_data = message.hex()
                        hex_array = [f"{int(x):02x}".upper() for x in message]
                        self.save_can_info(hex_array)
                        self.message_display('接收：' + ' '.join(f"{byte:02X}" for byte in message))
                        self.analysis_hex_data(hex_data)  # 解析数据
                    else:
                        buffer = buffer[1:]  # 丢弃无效数据头部
            except Exception as e:
                if self.running_ser:  # 仅在线程运行时显示错误
                    self.message_display('接收数据时出错：' + str(e), 'red')
                    self.message_display('================================')
                    messagebox.showerror('Error', '接收数据时出错：' + str(e))
                break

    '''发送线程'''

    def send_thread_func(self):
        while self.running_ser:
            try:
                if not self.send_queue.empty():
                    send_data = self.send_queue.get()
                    with self.serial_lock:  # 锁住串口线程，保证线程安全
                        self.ser.write(send_data)
                    # 转为16进制数组，CAN信息存储
                    hex_array = [f"{int(x):02x}".upper() for x in send_data]
                    self.save_can_info(hex_array)
                    send_data_str = ' '.join(i for i in hex_array)
                    self.message_display('发送：' + send_data_str)
                    self.send_queue.task_done()  # 任务完成，任务队列出队
            except Exception as e:
                if self.running_ser:  # 仅在线程运行时显示错误
                    self.message_display('发送数据时出错：' + str(e), 'red')
                    self.message_display('================================')
                    messagebox.showerror('Error', '发送数据时出错：' + str(e))
                break

    '''解析接收到的数据'''

    def analysis_hex_data(self, hex_data):
        if hex_data[:8] == 'aa010008':
            can_id = hex_data[8:16]
            data = hex_data[16:]
            '''根据CANID，解析数据'''
            match can_id:  # 流量数据，定义传过来的数据是计数值，需要进行转换为流量值
                case '00aa0401':
                    # hex_data[16:]中为8路流量数据，要先提取出来,成为8个元素的数组,拿到前6个元素
                    flow_data_array = [int(data[i:i + 2], 16) for i in range(0, len(data), 2)]
                    flow_values = [0, 0, 0, 0, 0, 0, 0, 0]
                    # 喷头处流量 公式有Q = 【F+10】/108
                    # 7516p = 1L 1/7516 * p
                    for i in range(len(flow_data_array) - 3):
                        flow_values[i] = (flow_data_array[i] + 10) / 108
                    # 管路流量1L=596脉冲 瞬时流量特性F=[10Q-4]
                    flow_values[5] = (flow_data_array[5] + 4) / 10
                    # 公式转换为流量 保留2位小数,放入self.flow_volumes进行显示
                    for i in range(len(flow_data_array) - 3):
                        self.flow_volumes[i].set(f'{flow_values[i]:.2f}')
                    self.channel_flow_var.set(f'{flow_values[5]:.2f}')
                case '00aa0501':  # 压力数据 获得值/100=实际值 预留位置
                    pressure_data = int(data[:2], 16) / 100
                    self.pressure_var.set(f'{pressure_data:.2f}')
                case _:  # 其他处理
                    pass

    '''打开/关闭串口'''

    def switch_serial_state(self):
        def button_permissions_open():
            self.port_choose['state'] = 'disabled'
            self.refresh_button['state'] = 'disabled'
            self.baudrate_choose['state'] = 'disabled'
            self.set_freq_button['state'] = 'normal'
            self.close_channel_button['state'] = 'normal'
            self.set_duty_phase_button['state'] = 'normal'

        def button_permissions_close():
            self.port_choose['state'] = 'normal'
            self.refresh_button['state'] = 'normal'
            self.baudrate_choose['state'] = 'normal'
            self.set_freq_button['state'] = 'disabled'
            self.close_channel_button['state'] = 'disabled'
            self.set_duty_phase_button['state'] = 'disabled'

        if self.serial_button['text'] == '打开':
            try:
                self.ser = serial.Serial(self.port_var.get(), int(self.baudrate_var.get()), timeout=0.5)
                self.serial_button['text'] = '关闭'
                # 按钮权限变更
                button_permissions_open()
                self.message_display('串口打开！串口号：' + self.port_var.get() + ', 波特率：' + self.baudrate_var.get(),
                                     'green')
                # 接收/发送线程打开
                self.running_ser = True
                # 接收线程打开
                self.receive_thread = threading.Thread(target=self.receive_thread_func, daemon=True)
                self.receive_thread.start()
                self.message_display('接收线程打开！', 'green')
                # 发送线程打开
                self.send_thread = threading.Thread(target=self.send_thread_func, daemon=True)
                self.send_thread.start()
                self.message_display('发送线程打开！', 'green')
                # excel后台线程打开
                # excel初始化
                self.excel_init()
                self.message_display('Excel文件初始化成功，文件名:' + self.xlsx_name, 'green')
                self.running_saver = True
                self.saver_thread = threading.Thread(target=self.saver_thread_func, daemon=True)
                self.saver_thread.start()
                self.message_display('Excel后台保存线程已启动', 'green')
                # 启动1S定时器
                self.one_second_timer()
                self.message_display('流量监测定时打开-1S', 'green')
                self.message_display('================================')
            except Exception as e:
                if '拒绝访问' in e.args[0]:
                    self.message_display('端口占用中,请断开端口连接!', 'red')
                    self.message_display('================================')
                    messagebox.showerror('Error', self.port_var.get() + '端口占用中,请断开端口连接')
                else:
                    self.message_display('未知错误：' + str(e), 'red')
                    self.message_display('================================')
                    messagebox.showerror('Error', '未知错误：' + str(e))
                return
        else:
            self.running_ser = False
            self.running_saver = False
            # 等待线程结束
            if self.receive_thread and self.receive_thread.is_alive():
                self.receive_thread.join(timeout=1)
                self.message_display('接收线程关闭！', 'red')
            if self.send_thread and self.send_thread.is_alive():
                self.send_thread.join(timeout=1)
                self.message_display('发送线程关闭！', 'red')
            if self.saver_thread and self.saver_thread.is_alive():
                self.saver_thread.join(timeout=1)
                self.message_display('Excel后台保存线程已停止!', 'red')
            # 确保串口不被占用时再关闭
            if self.ser and self.ser.is_open:
                self.ser.close()
            self.message_display('串口关闭！', 'red')
            self.message_display('================================')
            self.serial_button['text'] = '打开'
            # 按钮权限切换
            button_permissions_close()

    '''获取串口列表并显示'''

    def get_port_list(self):
        self.port_list = list(serial.tools.list_ports.comports())  # 获取串口列表
        if len(self.port_list) == 0:  # 不存在串口时，使用默认值
            self.port_list = ['COM1', 'COM2', 'COM3', 'COM4', 'COM5', 'COM6', 'COM7', 'COM8', 'COM9', 'COM10']
            self.port_var.set(self.port_list[8])
            self.message_display('不存在可用串口，使用默认选择项！', 'yellow')
        else:
            self.port_var.set(self.port_list[0].device)
            self.port_list = [self.port_list[i].device for i in range(len(self.port_list))]
            self.message_display('串口加载完成！', 'green')
        menu = self.port_choose['menu']
        menu.delete(0, 'end')
        for port in self.port_list:
            menu.add_command(label=port, command=lambda value=port: self.port_var.set(value))
        self.message_display('================================')

    '''设置频率'''

    def set_frequency(self, freq):
        # 发送十六进制信息，格式为aa010008+4字节CANID+8字节数据，例如AA0100081F0000000000000000000000
        # 将freq str转换为int，然后转换为2字节的16进制，并拆分为高低字节位，用于放入send_data的数组中
        # 当用户输入非数字时会导致程序崩溃，因此应该进行验证
        try:
            freq_val = 10 * int(freq)
        except ValueError:
            self.message_display("错误：请输入有效的数字", 'red')
            return
        self.message_display('设置电磁阀频率为：' + str(freq_val / 10) + 'Hz', 'green')
        freq_high_low = [(freq_val >> 8) & 0xFF, freq_val & 0xFF]
        send_data = [0xAA, 0x01, 0x00, 0x08, 0x1F, 0x00, 0x00, 0x00, freq_high_low[0], freq_high_low[1], 0x00, 0x00,
                     0x00, 0x00, 0x00, 0x00]
        self.send_queue.put(send_data)  # 将发送任务放入队列

    '''设置占空比和相位'''

    def set_duty_phase(self, channel, duty, phase):
        try:
            channel = int(channel)
            duty = 10 * int(duty)
            phase = 10 * int(phase)
        except ValueError:
            self.message_display("错误：请输入有效的数字", 'red')
            return
        self.message_display(
            '设置 ' + str(channel) + ' 号电磁阀：占空比：' + str(duty / 10) + '%，相位：' + str(phase / 10) + '°', 'green')
        duty_high_low = [(duty >> 8) & 0xFF, duty & 0xFF]
        phase_high_low = [(phase >> 8) & 0xFF, phase & 0xFF]
        send_data = [0xAA, 0x01, 0x00, 0x08, 0x1F, 0x00, 0x02, channel, duty_high_low[0], duty_high_low[1],
                     phase_high_low[0], phase_high_low[1], 0x00, 0x00, 0x00, 0x00]
        self.send_queue.put(send_data)  # 将发送任务放入队列

    '''关闭通道，打开另一个选择窗口'''

    def close_channel(self):
        self.close_channel_window = Toplevel(root)
        self.close_channel_window.title('关闭通道')
        # 一共5个通道，每个通道提供一个复选框
        for i in range(5):
            channel_var = StringVar()
            channel_var.set('喷头'+str(i + 1)+'号')
            channel_checkbutton = Checkbutton(self.close_channel_window, font=('黑体',13) ,textvariable=channel_var, variable=self.channels[i])
            channel_checkbutton.pack()
        close_channel_ok_button = Button(self.close_channel_window, font=('黑体',13) ,text='确定', command=self.close_channel_window_close)
        close_channel_ok_button.pack()

    def close_channel_window_close(self):
        self.close_channel_window.destroy()
        # 判断self.channels的是否选择上，如果选择上发送关闭指令入队列
        if self.channels[0].get():
            send_data = [0xAA, 0x01, 0x00, 0x08, 0x1F, 0x00, 0x02, 0x01, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00]
            self.send_queue.put(send_data)  # 将关闭任务放入队列
        for i in range(1,5):
            if self.channels[i].get():
                send_data = [0xAA, 0x01, 0x00, 0x08, 0x1F, 0x00, 0x02, i + 2, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00]
                self.send_queue.put(send_data)  # 将关闭任务放入队列

    '''读取脉冲计数值'''

    def read_flow_values(self):
        if not (self.running_ser and self.auto_read_flow_flag):
            return
        self.one_second_timer()
        send_data = [0xAA, 0x01, 0x00, 0x08, 0x00, 0xAA, 0x03, 0x01, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00]
        self.send_queue.put(send_data)

    '''1S 定时器'''

    def one_second_timer(self):
        # 定时器可能堆积，导致多个并发定时器运行,在启动新定时器前取消旧的
        if self.timer is not None and self.timer.is_alive():
            self.timer.cancel()
        self.timer = threading.Timer(1, self.read_flow_values)
        self.timer.start()

    '''文本框显示'''

    def message_display(self, message, color=None):
        # 向队列添加消息
        self.message_queue.put((message, color))
        if not self.after_id:
            self.after_id = self.after(self.MESSAGE_UPDATE_TIME, self.process_message_queue)

    '''信息队列处理'''

    def process_message_queue(self):
        batch_size = 20  # 每次最多处理20条
        self.message_box.config(state='normal')  # 使文本框可编辑
        try:
            for _ in range(batch_size):
                try:
                    msg, color = self.message_queue.get_nowait()
                    if color is None:
                        color = 'black'
                    self.message_box.tag_config(color, foreground=color)
                    self.message_box.insert(END, f'★{datetime.datetime.now()}: {msg}\n', color)
                except queue.Empty:
                    break
            self.message_box.see(END)  # 使文本框始终在最下面
        finally:
            self.message_box.config(state='disabled')  # 禁止编辑
            self.after_id = None if self.message_queue.empty() else self.after(self.MESSAGE_UPDATE_TIME,
                                                                               self.process_message_queue)

    '''切换自动发送流量flag'''

    def toggle_auto_read_flow(self):
        self.auto_read_flow_flag = not self.auto_read_flow_flag
        if self.auto_read_flow_flag:
            self.message_display('主动发送流量读取指令已开启！', 'green')
        else:
            self.message_display('主动发送流量读取指令已关闭！', 'red')
        self.read_flow_values()

    '''创建组件'''

    def create_widget(self):

        """框架区"""
        # 标题框架
        title_f = Frame(root, width=self.window_width, borderwidth=1, relief='solid', padx=30, pady=10)
        title_f.pack()
        # 配置区框架
        config_f = Frame(root, width=self.window_width, padx=30)
        config_serial_f = Frame(config_f)
        config_f.pack()
        config_serial_f.pack(pady=5, padx=30, side='bottom', anchor='w')
        # 传感器区框架
        sensor_data_f = Frame(root, width=self.window_width, padx=30)
        data_labels_f = Frame(sensor_data_f)
        sensor_data_f.pack()
        data_labels_f.pack(pady=5, padx=30, side='bottom', anchor='w')
        # 指令区框架
        order_f = Frame(root, width=self.window_width, padx=30)
        order_freq_f = Frame(order_f)
        order_duty_phase_f = Frame(order_f)
        order_f.pack()
        order_duty_phase_f.pack(pady=5, padx=30, side='bottom', anchor='w')
        order_freq_f.pack(pady=5, padx=30, side='bottom', anchor='w')
        # 信息区框架
        message_f = Frame(root, width=self.window_width, padx=30)
        message_f.pack()
        # 状态区框架
        bottom_f = Frame(root, width=self.window_width, padx=30)
        bottom_f.pack(side='bottom', anchor='w')

        '''内容区'''
        # 标题
        title_label = Label(title_f, text='PWM变量喷雾控制上位机软件', font=('黑体', 25))
        title_label.pack()

        # 配置区，用于配置串口
        # 分割线
        config_label = Label(config_f,
                             text='配置---------------------------------------------------------------------------',
                             font=('黑体', 15))
        config_label.pack()
        # 创建串口选择框、波特率选择框、打开/关闭按钮以及刷新按钮
        # 串口选择框
        label_port = Label(config_serial_f, text='串口：', font=('黑体', 15))
        label_port.pack(side='left')
        self.port_choose = OptionMenu(config_serial_f, self.port_var, *self.port_list)
        self.port_choose.pack(side='left')
        Label(config_serial_f, text='').pack(side='left', padx=10)
        # 波特率选择框
        label_baudrate = Label(config_serial_f, text='波特率：', font=('黑体', 15))
        label_baudrate.pack(side='left')

        self.baudrate_var.set(self.baudrate_list[0])
        self.baudrate_choose = OptionMenu(config_serial_f, self.baudrate_var, *self.baudrate_list)
        self.baudrate_choose.pack(side='left')
        # 串口开关按钮
        self.serial_button = Button(config_serial_f, text='打开', font=('黑体', 13),
                                    command=self.switch_serial_state)
        self.serial_button.pack(side='left', padx=30)
        # 刷新串口按钮
        self.refresh_button = Button(config_serial_f, text='刷新串口', font=('黑体', 13),
                                     command=self.get_port_list)
        self.refresh_button.pack(side='left')

        # 传感器数据区，观测流量、压力数据
        # 分割线
        flow_data_label = Label(sensor_data_f,
                                text='传感器数据---------------------------------------------------------------------',
                                font=('黑体', 15))
        flow_data_label.pack()
        # 创建2行3列 5个喷头流量，1个管路流量，1个管路压力
        # 流量5路
        flow_labels = [Label(data_labels_f, font=('黑体', 15), text='流量{}:'.format(i + 1)) for i in range(5)]
        self.flow_volumes = [StringVar() for _ in range(5)]
        flow_volume_labels = [Label(data_labels_f, font=('黑体', 15), width=5, bg='white', relief='solid',
                                    textvariable=self.flow_volumes[i]) for i in range(5)]
        flow_units = [Label(data_labels_f, font=('黑体', 15), text='L/min') for _ in range(5)]
        # 管路流量
        channel_flow_label = Label(data_labels_f, font=('黑体', 15), text='管路流量:')
        channel_flow_data_label = Label(data_labels_f, font=('黑体', 15), textvariable=self.channel_flow_var,
                                        bg='white', relief='solid', width=5)
        channel_flow_unit = Label(data_labels_f, font=('黑体', 15), text='L/min')
        # 管路压力
        pressure_label = Label(data_labels_f, font=('黑体', 15), text='管路压力:')

        pressure_data_label = Label(data_labels_f, font=('黑体', 15), textvariable=self.pressure_var, bg='white',
                                    relief='solid', width=5)
        pressure_unit = Label(data_labels_f, font=('黑体', 15), text='MPa  ')
        # 网格配置
        for i in range(5):
            flow_labels[i].grid(row=i // 3, column=i % 3 * 3)
            flow_volume_labels[i].grid(row=i // 3, column=i % 3 * 3 + 1)
            flow_units[i].grid(row=i // 3, column=i % 3 * 3 + 2, padx=3, pady=5)
            self.flow_volumes[i].set('0.00')
        channel_flow_label.grid(row=2, column=0)
        channel_flow_data_label.grid(row=2, column=1)
        self.channel_flow_var.set('0.00')
        channel_flow_unit.grid(row=2, column=2, pady=5)
        pressure_label.grid(row=2, column=3)
        pressure_data_label.grid(row=2, column=4)
        self.pressure_var.set('0.00')
        pressure_unit.grid(row=2, column=5, pady=5)

        # 指令发送区，差相驱动板指令发送
        # 分割线
        order_divider_label = Label(order_f,
                                    text='指令发送区---------------------------------------------------------------------',
                                    font=('黑体', 15))
        order_divider_label.pack()
        # 频率
        freq_label = Label(order_freq_f, text='频率:', font=('黑体', 15))
        freq_var = StringVar()
        freq_entry = Entry(order_freq_f, font=('黑体', 15), textvariable=freq_var, bg='white', relief='solid',
                           width=5, justify='center')
        freq_unit = Label(order_freq_f, text='Hz', font=('黑体', 15))
        freq_label.pack(side='left')
        freq_entry.pack(side='left')
        freq_unit.pack(side='left', padx=5)
        freq_var.set('1')
        # 设置频率按钮
        self.set_freq_button = Button(order_freq_f, state='disabled', text='频率设置', font=('黑体', 13),
                                      command=lambda: self.set_frequency(freq_entry.get()))
        self.set_freq_button.pack(side='left', padx=10)
        # 自动发送流量请求复选框
        auto_send_flow_var = IntVar()
        auto_send_flow_checkbutton = Checkbutton(order_freq_f, font=('黑体', 15), variable=auto_send_flow_var,
                                                 text='主动发送流量读取指令', onvalue=1, offvalue=0,
                                                 command=self.toggle_auto_read_flow)
        auto_send_flow_checkbutton.pack(side='left', padx=5)
        auto_send_flow_var.set(1)
        # 关闭通道
        self.close_channel_button = Button(order_freq_f, state='disabled', text='关闭通道', font=('黑体', 13),
                                           command=self.close_channel)
        self.close_channel_button.pack(side='left', padx=10)
        # 通道号
        duty_phase_channel_label = Label(order_duty_phase_f, text='通道:', font=('黑体', 15))
        duty_phase_channel_var = StringVar()
        duty_phase_channel_entry = Entry(order_duty_phase_f, font=('黑体', 15),
                                         textvariable=duty_phase_channel_var, bg='white', relief='solid',
                                         width=5, justify='center')
        duty_phase_channel_unit = Label(order_duty_phase_f, text='号', font=('黑体', 15))
        duty_phase_channel_label.pack(side='left')
        duty_phase_channel_entry.pack(side='left')
        duty_phase_channel_unit.pack(side='left', padx=5)
        duty_phase_channel_var.set('1')
        # 占空比
        duty_label = Label(order_duty_phase_f, text='占空比:', font=('黑体', 15))
        duty_var = StringVar()
        duty_entry = Entry(order_duty_phase_f, font=('黑体', 15), textvariable=duty_var, bg='white',
                           relief='solid', width=5, justify='center')
        duty_unit = Label(order_duty_phase_f, text='%', font=('黑体', 15))
        duty_label.pack(side='left')
        duty_entry.pack(side='left')
        duty_unit.pack(side='left', padx=5)
        duty_var.set('50')
        # 相位
        phase_label = Label(order_duty_phase_f, text='相位:', font=('黑体', 15))
        phase_var = StringVar()
        phase_entry = Entry(order_duty_phase_f, font=('黑体', 15), textvariable=phase_var, bg='white',
                            relief='solid', width=5, justify='center')
        phase_unit = Label(order_duty_phase_f, text='°', font=('黑体', 15))
        phase_label.pack(side='left')
        phase_entry.pack(side='left')
        phase_unit.pack(side='left', padx=5)
        phase_var.set('0')
        # 设置占空比和相位
        self.set_duty_phase_button = Button(order_duty_phase_f, state='disabled', text='占空比和相位设置',
                                            font=('黑体', 13),
                                            command=lambda: self.set_duty_phase(duty_phase_channel_var.get(),
                                                                                duty_var.get(),
                                                                                phase_var.get()))
        self.set_duty_phase_button.pack(side='left', padx=10)

        # 系统状态信息区
        # 分割线
        message_divider_label = Label(message_f,
                                      text='系统状态信息-------------------------------------------------------------------',
                                      font=('黑体', 15))
        message_divider_label.pack()
        # 信息框
        self.message_box = ScrolledText(message_f, width=100, height=8, font=('宋体', 10))
        self.message_box.pack(side='left', padx=30, pady=5)
        self.message_box.config(state='disabled')  # 禁止编辑

        # 状态栏
        version_label = Label(bottom_f, text='Version:', font=('黑体', 15))
        version_label.pack(side='left')
        current_version = get_local_version()
        version_currently = Label(bottom_f, text=current_version, font=('黑体', 15))
        version_currently.pack(side='left')

        # 显示信息、数据处理
        self.message_display('界面加载完成！', 'green')
        self.message_display('================================')
        self.get_port_list()


if __name__ == '__main__':
    root = Tk()
    root.geometry('900x600')  # 窗口大小
    root.title('PWM变量喷雾控制上位机软件')
    app = Concert(root)
    root.mainloop()
